import pandas as pd

import  os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

file_path = os.getcwd() + '\\financial_reports.xlsx'

def quartelyProfits(financialdata):
    df = pd.read_excel(financialdata)
    filtered = df[df['Year'] == 2014]
    print(filtered)
    ### Pivot Tables
    quarterly_profits = pd.pivot_table(filtered, index=df['Date'].dt.quarter, columns='Country', values='Profit', aggfunc='sum')

    ### Creating an Excel Workbook

    quarterly_profits.to_excel(file_path, sheet_name='Quarterly Profits 2014', startrow=3)

def main():
    filefinance = input("Please enter the financial workbook -- include extension (.xlsx if you are unsure)\n")
    # finWB = load_workbook(filefinance)
    # print(finWB.sheetnames)
    quartelyProfits(filefinance)

    # Load workbook
    wb = load_workbook(file_path)
    sheet = wb['Quarterly Profits 2014']

    for col in range(sheet.min_column, sheet.max_column + 1):
        #dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, width=45)
        sheet.column_dimensions[get_column_letter(col)].auto_size = True
        # print(col)
    #sheet.column_dimensions = dim_holder

    for col in sheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 6) * 1.5
        sheet.column_dimensions[column].width = adjusted_width

    sheet.merge_cells('A1:F3')
    sheet['A1'] = "Quarterly Profits"
    sheet['A1'].style = 'Title'
    sheet['A1'].alignment = Alignment(horizontal='center')


    for i in range(5, 9):
        sheet[f'B{i}'].style = 'Currency'
        sheet[f'C{i}'].style = 'Currency'
        sheet[f'D{i}'].style = 'Currency'
        sheet[f'E{i}'].style = 'Currency'
        sheet[f'F{i}'].style = 'Currency'

    # Add a Bar Chart
    bar_chart = BarChart()
    data = Reference(sheet, min_col=2,max_col=6, min_row=4,max_row=8)
    categories = Reference(sheet, min_col=1, max_col=1, min_row=5, max_row=8)
    bar_chart.style = 26
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    sheet.add_chart(bar_chart, "B11")

    bar_chart.title = 'Profits per Quarter'
    # bar_chart.style = 3

    wb.save(filename=file_path)


main()
